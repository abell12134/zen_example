#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF处理器 - 完整单文件版本
整合了PDF处理、OCR识别、Excel操作和GUI界面等功能，方便单文件打包
支持命令行模式和GUI模式
"""

import os
import sys
import shutil
import argparse
import re
import io
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from abc import ABC, abstractmethod

# GUI相关导入
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    print("GUI功能不可用，仅支持命令行模式")

# 第三方库导入
try:
    import fitz  # PyMuPDF
except ImportError:
    print("请安装PyMuPDF库: pip install PyMuPDF")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.drawing.image import Image as ExcelImage
except ImportError:
    print("请安装openpyxl库: pip install openpyxl")
    sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("请安装Pillow库: pip install Pillow")
    sys.exit(1)

# OCR库导入（可选）
RAPIDOCR_AVAILABLE = True
try:
    from rapidocr_onnxruntime import RapidOCR
except Exception as e:
    print(f"RapidOCR导入失败: {e}")
    RAPIDOCR_AVAILABLE = False


# ============================================================================
# 基础OCR类定义
# ============================================================================

class BaseOCR(ABC):
    @abstractmethod
    def __init__(self, lang: str = "ch"):
        pass

    @abstractmethod
    def recognize(self, img_bytes: bytes) -> List[Dict]:
        """
        识别图片中的文字
        :param img_bytes: 图片字节数据
        :return: 识别结果，格式为[{text: str, box: list, score: float}, ...]
        """
        pass

    @abstractmethod
    def close(self):
        pass


class RapidOCRWrapper(BaseOCR):
    def __init__(self, lang: str = "ch"):
        if not RAPIDOCR_AVAILABLE:
            raise ImportError(
                "RapidOCR is not available. This may be due to packaging limitations or missing configuration files."
            )
        
        self.lang = lang
        try:
            # 在打包环境中，尝试不使用配置文件初始化
            if getattr(sys, 'frozen', False):
                # 打包环境，使用最简单的初始化方式
                self.ocr = RapidOCR()
            else:
                # 正常环境
                self.ocr = RapidOCR(lang=lang)
        except Exception as e:
            print(f"RapidOCR初始化失败: {e}")
            # 尝试使用默认配置初始化
            try:
                self.ocr = RapidOCR()
                print("使用默认配置初始化RapidOCR成功")
            except Exception as e2:
                print(f"RapidOCR默认初始化也失败: {e2}")
                raise ImportError(f"RapidOCR无法初始化: {e2}")

    def recognize(self, img_bytes: bytes) -> List[Dict]:
        try:
            img = Image.open(io.BytesIO(img_bytes))
            
            # 转换为RGB格式以确保OCR兼容性
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            result, _ = self.ocr(img)

            if not result:
                return []

            output = []
            for line in result:
                box, text, score = line
                output.append({
                    "text": text,
                    "box": [
                        [box[0][0], box[0][1]],
                        [box[1][0], box[1][1]],
                        [box[2][0], box[2][1]],
                        [box[3][0], box[3][1]]
                    ],
                    "score": float(score)
                })
            return output
        except Exception as e:
            print(f"RapidOCR识别错误: {e}")
            import traceback
            traceback.print_exc()
            return []

    def close(self):
        pass


# ============================================================================
# 段落解析器
# ============================================================================

class ParagraphParse:
    def __init__(self, get_info, set_end) -> None:
        self.get_info = get_info
        self.set_end = set_end

    def run(self, text_blocks: list):
        units = self._get_units(text_blocks, self.get_info)
        self._parse(units)
        return text_blocks

    def _get_units(self, text_blocks, get_info):
        units = []
        for tb in text_blocks:
            bbox, text = get_info(tb)
            units.append((bbox, (text[0], text[-1]), tb))
        return units

    def _parse(self, units):
        if not units:
            return
        # 从bbox坐标点计算边界框
        def get_bbox_bounds(bbox):
            xs = [point[0] for point in bbox]
            ys = [point[1] for point in bbox]
            return min(xs), min(ys), max(xs), max(ys)
        
        units.sort(key=lambda a: get_bbox_bounds(a[0])[1])  # 按top排序
        para_l, para_top, para_r, para_bottom = get_bbox_bounds(units[0][0])
        para_line_h = para_bottom - para_top
        para_line_s = None
        now_para = [units[0]]
        paras = [now_para]
        paras_line_space = [para_line_s]

        for i in range(1, len(units)):
            l, top, r, bottom = get_bbox_bounds(units[i][0])
            h = bottom - top
            ls = top - para_bottom

            if (abs(para_l - l) <= para_line_h * 1.2 and
                abs(para_r - r) <= para_line_h * 1.2 and
                (para_line_s is None or ls < para_line_s + para_line_h * 0.5)):
                para_l = (para_l + l) / 2
                para_r = (para_r + r) / 2
                para_line_h = (para_line_h + h) / 2
                para_line_s = ls if para_line_s is None else (para_line_s + ls) / 2
                now_para.append(units[i])
            else:
                now_para = [units[i]]
                paras.append(now_para)
                paras_line_space.append(para_line_s)
                para_l, para_r, para_line_h = l, r, h
                para_line_s = None
            para_bottom = bottom

        for para in paras:
            for i1 in range(len(para) - 1):
                letter1 = para[i1][1][1]
                letter2 = para[i1 + 1][1][0]
                sep = self._word_separator(letter1, letter2)
                self.set_end(para[i1][2], sep)
            self.set_end(para[-1][2], "\n")

    def _word_separator(self, letter1, letter2):
        def is_cjk(character):
            cjk_ranges = [
                (0x4E00, 0x9FFF), (0x3040, 0x30FF), (0x1100, 0x11FF),
                (0x3130, 0x318F), (0xAC00, 0xD7AF), (0x3000, 0x303F),
                (0xFE30, 0xFE4F), (0xFF00, 0xFFEF)
            ]
            return any(start <= ord(character) <= end for start, end in cjk_ranges)

        if is_cjk(letter1) and is_cjk(letter2):
            return ""
        if letter1 == "-":
            return ""
        import unicodedata
        if unicodedata.category(letter2).startswith("P"):
            return ""
        return " "


# ============================================================================
# PDF处理器
# ============================================================================

class PDFProcessor:
    def __init__(self, ocr_engine: BaseOCR, password: str = ""):
        self.ocr_engine = ocr_engine
        self.password = password
        self.doc = None
        self.paragraph_parser = ParagraphParse(
            get_info=lambda tb: (tb["box"], tb["text"]),
            set_end=lambda tb, sep: tb.update({"end": sep})
        )

    def load_document(self, path: str) -> bool:
        try:
            self.doc = fitz.open(path)
            if self.doc.is_encrypted and not self.doc.authenticate(self.password):
                raise Exception("密码错误或文档已加密")
            return True
        except Exception as e:
            print(f"加载文档失败: {e}")
            return False

    def extract_images(self, page_num: int) -> List[Dict]:
        if not self.doc or page_num < 0 or page_num >= len(self.doc):
            return []

        page = self.doc[page_num]
        images = []
        img_list = page.get_images(full=True)

        for img in img_list:
            xref = img[0]
            base_image = self.doc.extract_image(xref)
            img_bytes = base_image["image"]
            img_rect = page.get_image_rects(xref)[0]

            images.append({
                "bytes": img_bytes,
                "bbox": [img_rect.x0, img_rect.y0, img_rect.x1, img_rect.y1],
                "width": base_image["width"],
                "height": base_image["height"]
            })
        return images

    def extract_text(self, page_num: int, mode: str = "mixed", ignore_areas: List[List[float]] = None) -> List[Dict]:
        """
        提取PDF页面文本
        :param page_num: 页码
        :param mode: 提取模式: mixed/textOnly/imageOnly/fullPage
        :param ignore_areas: 忽略区域 [[x0,y0,x1,y1], ...]
        :return: 文本块列表
        """
        if not self.doc or page_num < 0 or page_num >= len(self.doc):
            return []

        page = self.doc[page_num]
        text_blocks = []
        ignore_areas = ignore_areas or []

        # 提取原生文本
        if mode in ["mixed", "textOnly"]:
            blocks = page.get_text("dict")["blocks"]
            for block in blocks:
                if block["type"] == 0:  # 文本块
                    for line in block["lines"]:
                        text = "".join([span["text"] for span in line["spans"]])
                        if text and not self._is_in_ignore_area(line["bbox"], ignore_areas):
                            bbox = line["bbox"]
                            text_blocks.append({
                                "box": [
                                    [bbox[0], bbox[1]], [bbox[2], bbox[1]],
                                    [bbox[2], bbox[3]], [bbox[0], bbox[3]]
                                ],
                                "text": text,
                                "from": "text",
                                "end": ""
                            })

        # 处理图片中的文本
        if mode in ["mixed", "imageOnly", "fullPage"]:
            if mode == "fullPage":
                pix = page.get_pixmap()
                img_bytes = pix.tobytes("png")
                ocr_blocks = self._ocr_image(img_bytes)
                text_blocks.extend(self._filter_ignore_areas(ocr_blocks, ignore_areas))
            else:
                for img in self.extract_images(page_num):
                    ocr_blocks = self._ocr_image(img["bytes"])
                    text_blocks.extend(self._filter_ignore_areas(ocr_blocks, ignore_areas))

        return self.paragraph_parser.run(text_blocks)

    def _ocr_image(self, img_bytes: bytes) -> List[Dict]:
        ocr_result = self.ocr_engine.recognize(img_bytes)
        return [{
            "box": item["box"],
            "text": item["text"],
            "from": "ocr",
            "end": "",
            "score": item["score"]
        } for item in ocr_result]

    def _is_in_ignore_area(self, bbox: List[float], ignore_areas: List[List[float]]) -> bool:
        """检查文本块是否在忽略区域内"""
        x0, y0, x1, y1 = bbox
        center_x, center_y = (x0 + x1) / 2, (y0 + y1) / 2

        for area in ignore_areas:
            ax0, ay0, ax1, ay1 = area
            if ax0 <= center_x <= ax1 and ay0 <= center_y <= ay1:
                return True
        return False

    def _filter_ignore_areas(self, blocks: List[Dict], ignore_areas: List[List[float]]) -> List[Dict]:
        """过滤掉忽略区域内的文本块"""
        filtered = []
        for block in blocks:
            bbox = [
                block["box"][0][0], block["box"][0][1],
                block["box"][2][0], block["box"][2][1]
            ]
            if not self._is_in_ignore_area(bbox, ignore_areas):
                filtered.append(block)
        return filtered

    def close(self):
        if self.doc:
            self.doc.close()


# ============================================================================
# PDF到Excel处理器
# ============================================================================

class PDFToExcelProcessor:
    """PDF数据提取并写入Excel处理器"""
    
    def __init__(self, ocr_engine='rapid', lang='ch'):
        """初始化处理器
        
        Args:
            ocr_engine: OCR引擎类型 ('rapid')
            lang: 识别语言
        """
        # 初始化OCR引擎
        if RAPIDOCR_AVAILABLE:
            self.ocr_engine = RapidOCRWrapper(lang=lang)
        else:
            raise ImportError("RapidOCR不可用")
        
        self.pdf_processor = PDFProcessor(self.ocr_engine)
        
    def extract_pdf_data(self, pdf_path: str) -> List[Dict]:
        """从PDF中提取文本数据
        
        Args:
            pdf_path: PDF文件路径
            
        Returns:
            提取的文本块列表
        """
        print(f"开始处理PDF文件: {pdf_path}")
        
        if not self.pdf_processor.load_document(pdf_path):
            raise Exception(f"无法加载PDF文件: {pdf_path}")
        
        all_text_blocks = []
        
        try:
            # 处理所有页面
            for page_num in range(len(self.pdf_processor.doc)):
                print(f"正在处理第{page_num + 1}页...")
                text_blocks = self.pdf_processor.extract_text(page_num, mode="mixed")
                
                if text_blocks:
                    # 为每个文本块添加页码信息
                    for block in text_blocks:
                        block['page'] = page_num + 1
                    all_text_blocks.extend(text_blocks)
                    
            print(f"PDF处理完成，共提取{len(all_text_blocks)}个文本块")
            return all_text_blocks
            
        finally:
            self.pdf_processor.close()
    
    def convert_pdf_pages_to_images(self, pdf_path: str, output_dir: str = None) -> List[str]:
        """将PDF页面转换为图片
        
        Args:
            pdf_path: PDF文件路径
            output_dir: 图片输出目录，如果为None则使用PDF文件所在目录
            
        Returns:
            生成的图片文件路径列表
        """
        if output_dir is None:
            output_dir = os.path.dirname(pdf_path)
        
        pdf_filename = os.path.splitext(os.path.basename(pdf_path))[0]
        image_paths = []
        
        try:
            # 打开PDF文档
            doc = fitz.open(pdf_path)
            
            print(f"开始转换PDF页面为图片，共{len(doc)}页...")
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # 设置缩放比例以获得高质量图片
                zoom = 2.0  # 缩放比例
                mat = fitz.Matrix(zoom, zoom)
                
                # 渲染页面为图片
                pix = page.get_pixmap(matrix=mat)
                
                # 生成图片文件名
                image_filename = f"{pdf_filename}_page_{page_num + 1}.png"
                image_path = os.path.join(output_dir, image_filename)
                
                # 保存图片
                pix.save(image_path)
                image_paths.append(image_path)
                
                print(f"已保存第{page_num + 1}页: {image_filename}")
            
            doc.close()
            print(f"PDF转图片完成，共生成{len(image_paths)}张图片")
            return image_paths
            
        except Exception as e:
            print(f"PDF转图片时出错: {e}")
            return []
    
    def process_pdf_to_excel(self, pdf_path: str, template_excel_path: str = None, output_dir: str = None) -> bool:
        """处理PDF并写入Excel
        
        Args:
            pdf_path: PDF文件路径
            template_excel_path: Excel模板文件路径
            output_dir: 输出目录
            
        Returns:
            处理是否成功
        """
        try:
            # 设置输出目录
            if output_dir is None:
                output_dir = os.path.join(os.getcwd(), "output")
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成输出Excel文件名
            pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
            if template_excel_path:
                template_name = os.path.splitext(os.path.basename(template_excel_path))[0]
                output_excel_path = os.path.join(output_dir, f"{template_name}-{pdf_name}.xlsx")
                # 复制模板文件
                shutil.copy2(template_excel_path, output_excel_path)
            else:
                output_excel_path = os.path.join(output_dir, f"{pdf_name}.xlsx")
                # 创建新的Excel文件
                wb = openpyxl.Workbook()
                wb.save(output_excel_path)
            
            # 提取PDF数据
            text_blocks = self.extract_pdf_data(pdf_path)
            
            # 转换PDF页面为图片
            image_paths = self.convert_pdf_pages_to_images(pdf_path, output_dir)
            
            # 插入图片到Excel（如果有图片）
            if image_paths:
                self._insert_images_to_excel(output_excel_path, image_paths)
            
            print(f"处理完成: {pdf_path} -> {output_excel_path}")
            return True
            
        except Exception as e:
            print(f"处理失败: {e}")
            return False
    
    def _insert_images_to_excel(self, excel_path: str, image_paths: List[str]):
        """将图片插入到Excel文件中"""
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            # 创建或获取图纸工作表
            if '图纸' not in wb.sheetnames:
                drawing_ws = wb.create_sheet('图纸')
            else:
                drawing_ws = wb['图纸']
            
            # 清除现有图片
            if hasattr(drawing_ws, '_images') and drawing_ws._images:
                drawing_ws._images.clear()
            
            # 插入图片
            start_row = 2
            start_col = 2
            
            for i, image_path in enumerate(image_paths):
                if os.path.exists(image_path):
                    try:
                        # 创建Excel图片对象
                        img = ExcelImage(image_path)
                        
                        # 调整图片大小
                        img.width = 600
                        img.height = 800
                        
                        # 计算图片位置
                        if len(image_paths) <= 2:
                            col_offset = i * 10
                            row_offset = 0
                        else:
                            col_offset = (i % 2) * 10
                            row_offset = (i // 2) * 50
                        
                        # 设置图片锚点位置
                        cell_position = drawing_ws.cell(row=start_row + row_offset, 
                                                       column=start_col + col_offset)
                        img.anchor = cell_position.coordinate
                        
                        # 添加图片到工作表
                        drawing_ws.add_image(img)
                        
                        print(f"  插入图片: {os.path.basename(image_path)}")
                        
                    except Exception as e:
                        print(f"  插入图片 {image_path} 时出错: {e}")
            
            # 保存文件
            wb.save(excel_path)
            print(f"  图片插入完成")
            
        except Exception as e:
            print(f"插入图片到Excel时出错: {e}")


# ============================================================================
# 图片验证和重新插入功能
# ============================================================================

def find_image_for_excel(excel_filename: str, pdf_dir: str) -> List[str]:
    """根据Excel文件名找到对应的PNG图片文件"""
    # 从Excel文件名中提取PDF基础名称
    # 例如: "浙江锦康FAI 报告-浙江斐凌工具有限公司-45-34-1080__PN0012023_ACO LABEL -20250902.xlsx"
    # 提取: "45-34-1080__PN0012023_ACO"
    
    # 使用正则表达式匹配PDF基础名称模式
    # 匹配类似 "45-34-1080__PN0012023_ACO" 的模式
    pattern = r'(\d{2}-\d{2}-\d{4}__PN\d+_ACO)'
    match = re.search(pattern, excel_filename)
    
    if match:
        pdf_base = match.group(1)
        # 查找对应的PNG文件
        png_filename = f"{pdf_base}_page_1.png"
        png_path = os.path.join(pdf_dir, png_filename)
        
        if os.path.exists(png_path):
            return [png_path]
        else:
            print(f"警告: 找不到图片文件 {png_path}")
    else:
        print(f"警告: 无法从文件名 {excel_filename} 中提取PDF基础名称")
    
    return []


def verify_and_reinsert_images(excel_path: str, image_paths: List[str]) -> bool:
    """验证Excel文件中的图片是否存在，如果不存在则重新插入"""
    try:
        print(f"\n检查文件: {os.path.basename(excel_path)}")
        
        # 重新打开Excel文件检查图片
        wb = openpyxl.load_workbook(excel_path)
        
        if '图纸' not in wb.sheetnames:
            print("  未找到'图纸'工作表，需要重新插入图片")
            needs_reinsert = True
        else:
            drawing_ws = wb['图纸']
            # 检查是否有图片
            has_images = hasattr(drawing_ws, '_images') and len(drawing_ws._images) > 0
            
            if not has_images:
                print("  '图纸'工作表中没有图片，需要重新插入")
                needs_reinsert = True
            else:
                print(f"  '图纸'工作表中已有 {len(drawing_ws._images)} 张图片")
                needs_reinsert = False
        
        wb.close()
        
        # 如果需要重新插入图片
        if needs_reinsert and image_paths:
            print("  开始重新插入图片...")
            
            # 重新打开文件进行编辑
            wb = openpyxl.load_workbook(excel_path)
            
            # 创建或获取图纸工作表
            if '图纸' not in wb.sheetnames:
                drawing_ws = wb.create_sheet('图纸')
            else:
                drawing_ws = wb['图纸']
            
            # 清除现有图片
            if hasattr(drawing_ws, '_images') and drawing_ws._images:
                drawing_ws._images.clear()
            
            # 插入图片
            start_row = 2
            start_col = 2
            
            for i, image_path in enumerate(image_paths):
                if os.path.exists(image_path):
                    try:
                        # 创建Excel图片对象
                        img = ExcelImage(image_path)
                        
                        # 调整图片大小
                        img.width = 600
                        img.height = 800
                        
                        # 计算图片位置
                        if len(image_paths) <= 2:
                            col_offset = i * 10
                            row_offset = 0
                        else:
                            col_offset = (i % 2) * 10
                            row_offset = (i // 2) * 50
                        
                        # 设置图片锚点位置
                        cell_position = drawing_ws.cell(row=start_row + row_offset, 
                                                       column=start_col + col_offset)
                        img.anchor = cell_position.coordinate
                        
                        # 添加图片到工作表
                        drawing_ws.add_image(img)
                        
                        print(f"    插入图片: {os.path.basename(image_path)}")
                        
                    except Exception as e:
                        print(f"    插入图片 {image_path} 时出错: {e}")
            
            # 保存文件
            wb.save(excel_path)
            print("  图片重新插入完成")
            
            # 再次验证
            wb = openpyxl.load_workbook(excel_path)
            if '图纸' in wb.sheetnames:
                drawing_ws = wb['图纸']
                if hasattr(drawing_ws, '_images') and len(drawing_ws._images) > 0:
                    print(f"  验证成功: 图片已正确插入 ({len(drawing_ws._images)} 张)")
                    wb.close()
                    return True
                else:
                    print("  验证失败: 图片插入后仍然不存在")
                    wb.close()
                    return False
            else:
                print("  验证失败: '图纸'工作表不存在")
                wb.close()
                return False
        
        return True
        
    except Exception as e:
        print(f"验证和重新插入图片时出错: {e}")
        return False


def batch_verify_and_reinsert(output_dir: str):
    """批量验证并重新插入图片"""
    print(f"开始批量验证和重新插入图片，目录: {output_dir}")
    
    if not os.path.exists(output_dir):
        print(f"输出目录不存在: {output_dir}")
        return
    
    # 获取所有Excel文件
    excel_files = [f for f in os.listdir(output_dir) if f.lower().endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print("未找到Excel文件")
        return
    
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    success_count = 0
    
    for excel_file in excel_files:
        excel_path = os.path.join(output_dir, excel_file)
        
        # 查找对应的图片文件
        image_paths = find_image_for_excel(excel_file, output_dir)
        
        if not image_paths:
            print(f"跳过 {excel_file}: 未找到对应的图片文件")
            continue
        
        # 检查图片文件是否存在
        existing_images = [img for img in image_paths if os.path.exists(img)]
        
        if not existing_images:
            print(f"跳过 {excel_file}: 对应的图片文件不存在")
            continue
        
        # 验证并重新插入图片
        if verify_and_reinsert_images(excel_path, existing_images):
            success_count += 1
    
    print(f"\n批量验证和重新插入完成！成功处理: {success_count}/{len(excel_files)}")


# ============================================================================
# GUI界面类
# ============================================================================

if GUI_AVAILABLE:
    class PDFProcessorGUI:
        def __init__(self, root):
            self.root = root
            self.root.title("PDF处理工具 - 完整版")
            self.root.geometry("900x700")
            self.root.resizable(True, True)
            
            # 设置样式
            style = ttk.Style()
            style.theme_use('clam')
            
            # 变量
            self.pdf_file_path = tk.StringVar()
            self.template_file_path = tk.StringVar()
            self.output_dir_path = tk.StringVar()
            self.processing_mode = tk.StringVar(value="single")
            
            # 设置默认输出目录
            self.output_dir_path.set(os.path.join(os.getcwd(), "output"))
            
            self.create_widgets()
            
        def create_widgets(self):
            # 主框架
            main_frame = ttk.Frame(self.root, padding="10")
            main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # 配置网格权重
            self.root.columnconfigure(0, weight=1)
            self.root.rowconfigure(0, weight=1)
            main_frame.columnconfigure(1, weight=1)
            
            # 标题
            title_label = ttk.Label(main_frame, text="PDF处理工具 - 完整版", font=('Arial', 16, 'bold'))
            title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
            
            # 处理模式选择
            mode_frame = ttk.LabelFrame(main_frame, text="处理模式", padding="10")
            mode_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
            
            ttk.Radiobutton(mode_frame, text="单文件处理", variable=self.processing_mode, 
                           value="single", command=self.on_mode_change).grid(row=0, column=0, padx=(0, 20))
            ttk.Radiobutton(mode_frame, text="批量处理", variable=self.processing_mode, 
                           value="batch", command=self.on_mode_change).grid(row=0, column=1)
            
            # 文件选择区域
            file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
            file_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
            file_frame.columnconfigure(1, weight=1)
            
            # PDF文件选择
            self.pdf_label = ttk.Label(file_frame, text="PDF文件:")
            self.pdf_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
            
            self.pdf_entry = ttk.Entry(file_frame, textvariable=self.pdf_file_path, width=50)
            self.pdf_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(0, 5))
            
            self.pdf_button = ttk.Button(file_frame, text="浏览", command=self.browse_pdf_file)
            self.pdf_button.grid(row=0, column=2, pady=(0, 5))
            
            # Excel模板文件选择
            ttk.Label(file_frame, text="Excel模板:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
            
            ttk.Entry(file_frame, textvariable=self.template_file_path, width=50).grid(
                row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(0, 5))
            
            ttk.Button(file_frame, text="浏览", command=self.browse_template_file).grid(
                row=1, column=2, pady=(0, 5))
            
            # 输出目录选择
            ttk.Label(file_frame, text="输出目录:").grid(row=2, column=0, sticky=tk.W)
            
            ttk.Entry(file_frame, textvariable=self.output_dir_path, width=50).grid(
                row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
            
            ttk.Button(file_frame, text="浏览", command=self.browse_output_dir).grid(
                row=2, column=2)
            
            # 操作按钮区域
            button_frame = ttk.Frame(main_frame)
            button_frame.grid(row=3, column=0, columnspan=3, pady=20)
            
            self.process_button = ttk.Button(button_frame, text="开始处理", 
                                           command=self.start_processing, style='Accent.TButton')
            self.process_button.grid(row=0, column=0, padx=(0, 10))
            
            self.verify_button = ttk.Button(button_frame, text="验证并重插入图片", 
                                          command=self.start_verification)
            self.verify_button.grid(row=0, column=1, padx=(0, 10))
            
            ttk.Button(button_frame, text="清空日志", command=self.clear_log).grid(row=0, column=2)
            
            # 进度条
            self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
            self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
            
            # 日志显示区域
            log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="10")
            log_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
            log_frame.columnconfigure(0, weight=1)
            log_frame.rowconfigure(0, weight=1)
            main_frame.rowconfigure(5, weight=1)
            
            self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80)
            self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # 初始化界面状态
            self.on_mode_change()
            
        def on_mode_change(self):
            """处理模式改变时的界面更新"""
            if self.processing_mode.get() == "batch":
                # 批量模式下禁用PDF文件选择
                self.pdf_entry.config(state='disabled')
                self.pdf_button.config(state='disabled')
                self.pdf_label.config(text="PDF目录: (批量模式)")
                self.log("切换到批量处理模式")
            else:
                # 单文件模式下启用PDF文件选择
                self.pdf_entry.config(state='normal')
                self.pdf_button.config(state='normal')
                self.pdf_label.config(text="PDF文件:")
                self.log("切换到单文件处理模式")
        
        def browse_pdf_file(self):
            """浏览PDF文件"""
            filename = filedialog.askopenfilename(
                title="选择PDF文件",
                filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
            )
            if filename:
                self.pdf_file_path.set(filename)
                self.log(f"选择PDF文件: {filename}")
        
        def browse_template_file(self):
            """浏览Excel模板文件"""
            filename = filedialog.askopenfilename(
                title="选择Excel模板文件",
                filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
            )
            if filename:
                self.template_file_path.set(filename)
                self.log(f"选择Excel模板: {filename}")
        
        def browse_output_dir(self):
            """浏览输出目录"""
            dirname = filedialog.askdirectory(title="选择输出目录")
            if dirname:
                self.output_dir_path.set(dirname)
                self.log(f"设置输出目录: {dirname}")
        
        def log(self, message):
            """添加日志信息"""
            self.log_text.insert(tk.END, f"{message}\n")
            self.log_text.see(tk.END)
            self.root.update_idletasks()
        
        def clear_log(self):
            """清空日志"""
            self.log_text.delete(1.0, tk.END)
        
        def validate_inputs(self):
            """验证输入参数"""
            if self.processing_mode.get() == "single":
                if not self.pdf_file_path.get():
                    messagebox.showerror("错误", "请选择PDF文件")
                    return False
                if not os.path.exists(self.pdf_file_path.get()):
                    messagebox.showerror("错误", "PDF文件不存在")
                    return False
            
            if not self.template_file_path.get():
                messagebox.showerror("错误", "请选择Excel模板文件")
                return False
            
            if not os.path.exists(self.template_file_path.get()):
                messagebox.showerror("错误", "Excel模板文件不存在")
                return False
            
            if not self.output_dir_path.get():
                messagebox.showerror("错误", "请设置输出目录")
                return False
            
            return True
        
        def start_processing(self):
            """开始处理"""
            if not self.validate_inputs():
                return
            
            # 禁用按钮并显示进度条
            self.process_button.config(state='disabled')
            self.progress.start()
            
            # 在新线程中执行处理
            thread = threading.Thread(target=self.process_files)
            thread.daemon = True
            thread.start()
        
        def process_files(self):
            """处理文件的主要逻辑"""
            try:
                output_dir = self.output_dir_path.get()
                template_path = self.template_file_path.get()
                
                # 确保输出目录存在
                os.makedirs(output_dir, exist_ok=True)
                
                if self.processing_mode.get() == "single":
                    # 单文件处理
                    pdf_path = self.pdf_file_path.get()
                    self.log(f"开始处理单个PDF文件: {pdf_path}")
                    
                    processor = PDFToExcelProcessor()
                    success = processor.process_pdf_to_excel(
                        pdf_path=pdf_path,
                        template_excel_path=template_path,
                        output_dir=output_dir
                    )
                    
                    if success:
                        self.log("PDF处理完成！")
                        messagebox.showinfo("成功", "PDF处理完成！")
                    else:
                        self.log("PDF处理失败！")
                        messagebox.showerror("错误", "PDF处理失败！")
                
                else:
                    # 批量处理
                    self.log("开始批量处理PDF文件...")
                    
                    # 查找pdf_process_ing目录
                    pdf_dir = os.path.join(os.getcwd(), "pdf_process_ing")
                    if not os.path.exists(pdf_dir):
                        self.log(f"未找到pdf_process_ing目录: {pdf_dir}")
                        messagebox.showerror("错误", f"未找到pdf_process_ing目录: {pdf_dir}")
                        return
                    
                    # 获取所有PDF文件
                    pdf_files = [f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf')]
                    if not pdf_files:
                        self.log("pdf_process_ing目录中没有找到PDF文件")
                        messagebox.showwarning("警告", "pdf_process_ing目录中没有找到PDF文件")
                        return
                    
                    self.log(f"找到 {len(pdf_files)} 个PDF文件")
                    
                    processor = PDFToExcelProcessor()
                    success_count = 0
                    
                    for i, pdf_file in enumerate(pdf_files, 1):
                        pdf_path = os.path.join(pdf_dir, pdf_file)
                        self.log(f"处理第 {i}/{len(pdf_files)} 个文件: {pdf_file}")
                        
                        try:
                            success = processor.process_pdf_to_excel(
                                pdf_path=pdf_path,
                                template_excel_path=template_path,
                                output_dir=output_dir
                            )
                            
                            if success:
                                success_count += 1
                                self.log(f"✓ {pdf_file} 处理成功")
                            else:
                                self.log(f"✗ {pdf_file} 处理失败")
                        
                        except Exception as e:
                            self.log(f"✗ {pdf_file} 处理出错: {str(e)}")
                    
                    self.log(f"批量处理完成！成功: {success_count}/{len(pdf_files)}")
                    messagebox.showinfo("完成", f"批量处理完成！\n成功: {success_count}/{len(pdf_files)}")
            
            except Exception as e:
                error_msg = f"处理过程中发生错误: {str(e)}"
                self.log(error_msg)
                messagebox.showerror("错误", error_msg)
            
            finally:
                # 恢复界面状态
                self.root.after(0, self.processing_finished)
        
        def start_verification(self):
            """开始验证并重插入图片"""
            if not self.output_dir_path.get():
                messagebox.showerror("错误", "请设置输出目录")
                return
            
            # 禁用按钮并显示进度条
            self.verify_button.config(state='disabled')
            self.progress.start()
            
            # 在新线程中执行验证
            thread = threading.Thread(target=self.verify_images)
            thread.daemon = True
            thread.start()
        
        def verify_images(self):
            """验证并重插入图片的主要逻辑"""
            try:
                output_dir = self.output_dir_path.get()
                self.log("开始验证并重插入图片...")
                
                # 调用批量验证函数
                batch_verify_and_reinsert(output_dir)
                
                self.log("图片验证和重插入完成！")
                messagebox.showinfo("成功", "图片验证和重插入完成！")
            
            except Exception as e:
                error_msg = f"验证过程中发生错误: {str(e)}"
                self.log(error_msg)
                messagebox.showerror("错误", error_msg)
            
            finally:
                # 恢复界面状态
                self.root.after(0, self.verification_finished)
        
        def processing_finished(self):
            """处理完成后的界面恢复"""
            self.progress.stop()
            self.process_button.config(state='normal')
        
        def verification_finished(self):
            """验证完成后的界面恢复"""
            self.progress.stop()
            self.verify_button.config(state='normal')


# ============================================================================
# 命令行处理函数
# ============================================================================

def process_pdf_to_excel(pdf_path: str, template_excel_path: str = None) -> bool:
    """处理单个PDF文件到Excel（兼容原有接口）"""
    try:
        processor = PDFToExcelProcessor()
        return processor.process_pdf_to_excel(pdf_path, template_excel_path)
    except Exception as e:
        print(f"处理PDF文件失败: {e}")
        return False


def process_all_pdfs_in_directory(pdf_directory: str, template_excel_path: str, output_dir: str = None) -> int:
    """批量处理目录中的所有PDF文件"""
    if not os.path.exists(pdf_directory):
        print(f"PDF目录不存在: {pdf_directory}")
        return 0
    
    if output_dir is None:
        output_dir = os.path.join(os.getcwd(), "output")
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 获取所有PDF文件
    pdf_files = [f for f in os.listdir(pdf_directory) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print(f"在目录 {pdf_directory} 中未找到PDF文件")
        return 0
    
    print(f"找到 {len(pdf_files)} 个PDF文件，开始批量处理...")
    
    processor = PDFToExcelProcessor()
    success_count = 0
    
    for i, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(pdf_directory, pdf_file)
        print(f"处理第 {i}/{len(pdf_files)} 个文件: {pdf_file}")
        
        try:
            success = processor.process_pdf_to_excel(
                pdf_path=pdf_path,
                template_excel_path=template_excel_path,
                output_dir=output_dir
            )
            
            if success:
                success_count += 1
                print(f"✓ {pdf_file} 处理成功")
            else:
                print(f"✗ {pdf_file} 处理失败")
        
        except Exception as e:
            print(f"✗ {pdf_file} 处理出错: {str(e)}")
    
    print(f"批量处理完成！成功: {success_count}/{len(pdf_files)}")
    return success_count


# ============================================================================
# 主函数
# ============================================================================

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="PDF处理工具 - 支持单个文件处理、批量处理和GUI模式",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""使用示例:
  单文件处理:
    python pdf_processor_complete.py -p /path/to/file.pdf -t /path/to/template.xlsx
    python pdf_processor_complete.py -p /path/to/file.pdf -t /path/to/template.xlsx -o /path/to/output
  
  批量处理:
    python pdf_processor_complete.py --batch -t /path/to/template.xlsx
    python pdf_processor_complete.py --batch -t /path/to/template.xlsx -o /path/to/output
  
  GUI模式:
    python pdf_processor_complete.py --gui
        """
    )
    
    parser.add_argument('-p', '--pdf', type=str, help='PDF文件路径')
    parser.add_argument('-t', '--template', type=str, help='Excel模板文件路径')
    parser.add_argument('-o', '--output', type=str, help='输出目录路径（可选，默认为当前目录下的output文件夹）')
    parser.add_argument('--batch', action='store_true', help='批量处理模式（处理pdf_process_ing目录下的所有PDF文件）')
    parser.add_argument('--gui', action='store_true', help='启动GUI界面模式')
    parser.add_argument('--verify', action='store_true', help='验证并重新插入图片模式')
    
    args = parser.parse_args()
    
    # GUI模式
    if args.gui:
        if not GUI_AVAILABLE:
            print("错误: GUI功能不可用，请安装tkinter库")
            sys.exit(1)
        
        root = tk.Tk()
        app = PDFProcessorGUI(root)
        
        # 居中显示窗口
        root.update_idletasks()
        width = root.winfo_width()
        height = root.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        root.geometry(f"{width}x{height}+{x}+{y}")
        
        root.mainloop()
        return
    
    # 验证模式
    if args.verify:
        output_dir = args.output or os.path.join(os.getcwd(), "output")
        batch_verify_and_reinsert(output_dir)
        return
    
    # 命令行模式
    if args.batch:
        # 批量处理模式
        if not args.template:
            print("错误: 批量处理模式需要指定Excel模板文件 (-t/--template)")
            sys.exit(1)
        
        pdf_dir = os.path.join(os.getcwd(), "pdf_process_ing")
        output_dir = args.output or os.path.join(os.getcwd(), "output")
        
        success_count = process_all_pdfs_in_directory(pdf_dir, args.template, output_dir)
        
        if success_count > 0:
            print(f"\n批量处理完成！成功处理 {success_count} 个文件")
        else:
            print("\n批量处理失败或没有文件被处理")
            sys.exit(1)
    
    else:
        # 单文件处理模式
        if not args.pdf:
            print("错误: 单文件处理模式需要指定PDF文件 (-p/--pdf)")
            print("使用 --help 查看帮助信息")
            sys.exit(1)
        
        if not args.template:
            print("错误: 需要指定Excel模板文件 (-t/--template)")
            sys.exit(1)
        
        # 设置输出目录
        if args.output:
            output_dir = args.output
        else:
            output_dir = os.path.join(os.getcwd(), "output")
        
        # 处理单个PDF文件
        processor = PDFToExcelProcessor()
        success = processor.process_pdf_to_excel(
            pdf_path=args.pdf,
            template_excel_path=args.template,
            output_dir=output_dir
        )
        
        if success:
            print("\nPDF处理完成！")
        else:
            print("\nPDF处理失败！")
            sys.exit(1)


if __name__ == "__main__":
    main()